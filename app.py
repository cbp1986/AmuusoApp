import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from io import BytesIO

def generate_offer(input_file, tarief, taal, include_stock, lego_mode, discount=None, uitloop=False, promo=False):
    datum = datetime.now().strftime("%Y%m%d")
    logo_file = "logo.png"
    
    try:
        df = pd.read_excel(input_file, dtype=str)
    except Exception as e:
        st.error(f"Fout bij het openen van het bestand: {str(e)}")
        return None
    
    if lego_mode:
        df = df.rename(columns={"artikelnummer": "Amuuso Ref", "barcode": "GTIN", "Statuscode": "Code"})
        df = df[["Amuuso Ref", "e-com omschr nl", "e-com omschr fr", "GTIN", "vkp tarief 4", "vkp tarief 5", "winkelprijs1", "voorraad"]]
        df = df.fillna("")
        df = df[df["voorraad"] != "0"]
    
    df["Prijs/Prix"] = df["vkp tarief 4"] if tarief == "T4" else df["vkp tarief 5"]
    df["SRP"] = df["winkelprijs1"]
    df["Stock"] = df["voorraad"] if include_stock else ""
    
    if discount and promo:
        df["Promo"] = df["Prijs/Prix"].astype(float) * (1 - discount / 100)
    
    if uitloop:
        df = df[df["rotlev"] == "U"]
    
    df = df[[col for col in ["Amuuso Ref", "e-com omschr nl", "e-com omschr fr", "GTIN", "Prijs/Prix", "SRP", "Stock", "Promo"] if col in df.columns]]
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=2)
        wb = writer.book
        ws = writer.sheets["Sheet1"]
        ws.title = f"Offerte {datum}"
        
        if os.path.exists(logo_file):
            img = Image(logo_file)
            ws.add_image(img, "A1")
        
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[3]:
            cell.font = Font(size=14, bold=True)
            cell.fill = header_fill
        
    output.seek(0)
    return output

st.title("Offerte Generator")

uploaded_file = st.file_uploader("Upload een Excel bestand", type=["xlsx", "xls"])

tarief = st.radio("Kies prijstarief:", ["T4", "T5"], index=0)
taal = st.radio("Kies taal:", ["NL", "FR"], index=0)
include_stock = st.checkbox("Inclusief voorraad?")
lego_mode = st.checkbox("LEGO Stocklijst?")
uitloop = st.checkbox("Enkel uitloop artikelen tonen?")
promo_active = st.checkbox("Promo korting toepassen?")
discount = st.number_input("Korting in %", min_value=0.0, max_value=100.0, step=0.1) if promo_active else None

if uploaded_file and st.button("Offerte genereren"):
    output = generate_offer(uploaded_file, tarief, taal, include_stock, lego_mode, discount, uitloop, promo_active)
    if output:
        st.download_button(label="Download Excel bestand", data=output, file_name=f"Offerte_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
